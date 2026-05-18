#!/usr/bin/env python3
import sys, json, re, os
import openpyxl
data = json.loads(sys.stdin.read())
orders = data['orders']
with open(os.path.join(os.path.dirname(__file__), 'localidades_andreani.json')) as f:
    localidades = json.load(f)
PROVINCIAS = {
  'A': 'SALTA', 'B': 'BUENOS AIRES', 'C': 'CIUDAD AUTONOMA DE BUENOS AIRES',
  'D': 'SAN LUIS', 'E': 'ENTRE RIOS', 'F': 'LA RIOJA', 'G': 'SANTIAGO DEL ESTERO',
  'H': 'CHACO', 'J': 'SAN JUAN', 'K': 'CATAMARCA', 'L': 'LA PAMPA', 'M': 'MENDOZA',
  'N': 'MISIONES', 'P': 'FORMOSA', 'Q': 'NEUQUEN', 'R': 'RIO NEGRO', 'S': 'SANTA FE',
  'T': 'TUCUMAN', 'U': 'CHUBUT', 'V': 'TIERRA DEL FUEGO', 'W': 'CORRIENTES',
  'X': 'CORDOBA', 'Y': 'JUJUY', 'Z': 'SANTA CRUZ'
}
loc_map = {}
for l in localidades:
    parts = l.split(' / ')
    if len(parts) == 3:
        prov, loc, cp = parts
        cpNum = re.sub(r'[^0-9]', '', cp)
        loc_map[f'{prov}|{loc}|{cpNum}'] = l
        if f'{prov}|{cpNum}' not in loc_map:
            loc_map[f'{prov}|{cpNum}'] = l
def buscar_loc(prov_code, ciudad, zip_):
    prov = PROVINCIAS.get(prov_code.upper(), prov_code.upper())
    loc = ciudad.upper().strip()
    cpNum = re.sub(r'[^0-9]', '', zip_)[:4]
    if loc_map.get(f'{prov}|{loc}|{cpNum}'): return loc_map[f'{prov}|{loc}|{cpNum}']
    if loc_map.get(f'{prov}|{cpNum}'): return loc_map[f'{prov}|{cpNum}']
    found = next((l for l in localidades if re.sub(r'[^0-9]', '', l.split(' / ')[2] if len(l.split(' / '))>2 else '') == cpNum), None)
    return found or f'{prov} / {loc} / {cpNum}'
def parse_tel(raw):
    if not raw: return '', ''
    digits = re.sub(r'\D', '', str(raw))
    if digits.startswith('549'): digits = digits[3:]
    elif digits.startswith('54'): digits = digits[2:]
    if digits.startswith('0'): digits = digits[1:]
    if len(digits) == 10:
        if digits[:2] in ['11','22','23','26','27','28','29','30']:
            return digits[:2], digits[2:]
        else:
            return digits[:3], digits[3:]
    elif len(digits) == 8: return '11', digits
    elif len(digits) == 7: return digits[:3], digits[3:]
    return digits[:3], digits[3:]
def parse_addr(a1, a2):
    a1 = a1 or ''; a2 = a2 or ''
    m = re.match(r'^(.*?)\s+(\d+[a-zA-Z]?)(?:\s+(.*))?$', a1)
    if m:
        calle = m.group(1).strip()
        num = m.group(2).strip()
        resto = m.group(3) or ''
        piso = re.search(r'[Pp]iso\s*(\w+)', resto) or re.search(r'[Pp]iso\s*(\w+)', a2)
        depto = re.search(r'[Dd]pto\.?\s*([A-Za-z0-9]+)', resto) or re.search(r'[Dd]pto\.?\s*([A-Za-z0-9]+)', a2)
        return calle, num, piso.group(1) if piso else '', depto.group(1) if depto else ''
    return a1.strip(), '0', '', ''
wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), 'plantilla_andreani (1).xlsx'))
ws = wb['A domicilio']
for idx, o in enumerate(orders):
    row = idx + 3
    addr = o.get('shipping_address') or o.get('billing_address') or {}
    calle, numero, piso, depto = parse_addr(addr.get('address1',''), addr.get('address2',''))
    if not numero: numero = '0'
    cod_tel, num_tel = parse_tel(addr.get('phone') or o.get('phone',''))
    loc_val = buscar_loc(addr.get('province_code',''), addr.get('city',''), addr.get('zip',''))
    customer = o.get('customer') or {}
    ws.cell(row, 1).value  = 'PAQUETE'
    ws.cell(row, 2).value  = None
    ws.cell(row, 3).value  = None
    ws.cell(row, 4).value  = None
    ws.cell(row, 5).value  = None
try:
    ws.cell(row, 6).value = float(str(o.get('total_price') or '0').replace(',', '.').replace(' ', '') or '0')
except:
    ws.cell(row, 6).value = 0.0
    ws.cell(row, 7).value  = str(o.get('order_number','')).replace('#','')
    ws.cell(row, 8).value  = addr.get('first_name') or customer.get('first_name','')
    ws.cell(row, 9).value  = addr.get('last_name') or customer.get('last_name','')
    ws.cell(row, 10).value = '00000000'
    ws.cell(row, 11).value = o.get('email') or customer.get('email','')
    ws.cell(row, 12).value = cod_tel
    ws.cell(row, 13).value = num_tel
    ws.cell(row, 14).value = calle
    ws.cell(row, 15).value = numero
    ws.cell(row, 16).value = piso
    ws.cell(row, 17).value = depto
    ws.cell(row, 18).value = loc_val
    ws.cell(row, 19).value = ''
wb.save(sys.stdout.buffer)
